from __future__ import annotations

import csv
import json
import re
from pathlib import Path
from typing import Any, Dict, List, Optional

import openpyxl


PARSER_VERSION = 5
SHEET_ALIASES = {
    1: "周擂台1",
    2: "周擂台2",
}
FIXED_LUNATIC_HP = {
    1: {1: 67499, 2: 74999, 3: 82499},
    2: {1: 33749, 2: 37499, 3: 41249},
    3: {1: 22499, 2: 24999, 3: 27499},
}
ELEMENT_CHAR_TO_ID = {"日": 1, "月": 2, "火": 3, "水": 4, "木": 5, "金": 6, "土": 7, "星": 8}
STAT_SUB_IDS = {
    "阳攻": 1,
    "阳防": 2,
    "阴攻": 3,
    "阴防": 4,
    "速度": 5,
    "命中": 6,
    "回避": 7,
    "会心攻击": 8,
    "会心命中": 9,
}
ANOMALY_SUB_IDS = {"燃烧": 1, "冻结": 2, "感电": 3, "毒雾": 4, "黑暗": 5}
BULLET_SUB_IDS = {
    "通常弹": 1,
    "镭射弹": 3,
    "体术弹": 4,
    "斩击弹": 5,
    "动能弹": 6,
    "流体弹": 7,
    "能量弹": 8,
    "御符弹": 9,
    "光弹": 10,
    "尖弹": 11,
    "追踪弹": 12,
}


def find_arena_xlsx(base_dir: Path) -> Optional[Path]:
    candidates = [
        base_dir.parent / "擂台敌人数据详情.xlsx",
        base_dir / "擂台敌人数据详情.xlsx",
        Path.cwd() / "擂台敌人数据详情.xlsx",
    ]
    return next((path for path in candidates if path.exists()), None)


def _to_int(value: Any, default: int = 0) -> int:
    try:
        if value is None or value == "":
            return default
        return int(float(value))
    except Exception:
        return default


def _to_float(value: Any, default: float = 0.0) -> float:
    try:
        if value is None or value == "":
            return default
        return float(value)
    except Exception:
        return default


def _text(value: Any) -> str:
    return str(value or "").strip()


def _quality_from_text(weak_text: str, resist_text: str) -> List[int]:
    quality = [1] * 9
    for ch in str(weak_text or ""):
        idx = ELEMENT_CHAR_TO_ID.get(ch)
        if idx:
            quality[idx - 1] = 0
    for ch in str(resist_text or ""):
        idx = ELEMENT_CHAR_TO_ID.get(ch)
        if idx:
            quality[idx - 1] = 2
    quality[8] = 1
    return quality


def _duration_from_text(text: str, default: int = 1) -> int:
    match = re.search(r"\((\d+)\s*回合\)|（(\d+)\s*回合）", text)
    if not match:
        return default
    return _to_int(match.group(1) or match.group(2), default)


def parse_effect_text(text: str) -> List[List[int]]:
    text = _text(text)
    if not text:
        return []
    duration = _duration_from_text(text, 1)
    effects: List[List[int]] = []
    target = 2 if "己方全体" in text else 1 if "自身" in text else 4 if "敌方全体" in text else 3 if "敌方单体" in text else 2

    for stat, sub_id in STAT_SUB_IDS.items():
        if stat in text:
            match = re.search(rf"{re.escape(stat)}(?:Ⅱ)?(?:上升|下降)(\d+)等级", text)
            if match:
                effects.append([1 if "上升" in match.group(0) else 2, sub_id, target, duration, _to_int(match.group(1), 0)])

    for anomaly, sub_id in ANOMALY_SUB_IDS.items():
        if anomaly in text and ("附加" in text or "付与" in text):
            match = re.search(r"附加(\d+)枚|付与(\d+)枚", text)
            effects.append([6, sub_id, target, duration, _to_int((match.group(1) or match.group(2)) if match else 1, 1)])

    if "结界回复" in text or "结界增加" in text:
        match = re.search(r"(?:结界回复|结界增加)(\d+)枚", text)
        effects.append([4, 0, target, duration, _to_int(match.group(1) if match else 1, 1)])

    if "灵力上升" in text:
        match = re.search(r"灵力上升([0-9.]+)", text)
        spirit = _to_float(match.group(1) if match else 1.0, 1.0)
        effects.append([5, 0, target, duration, int(round(spirit * 20))])

    for bullet, sub_id in BULLET_SUB_IDS.items():
        if bullet in text and "伤害下降" in text:
            match = re.search(r"伤害下降(\d+)%", text)
            effects.append([12, sub_id, target, duration, _to_int(match.group(1) if match else 0, 0)])

    return effects


def _parse_base_character_id(name: str) -> int:
    matches = re.findall(r"\((\d+)\)", str(name or ""))
    return _to_int(matches[-1], 0) if matches else 0


def _cell(rows: List[tuple], row: int, col: int) -> Any:
    r = row - 1
    c = col - 1
    if r < 0 or r >= len(rows):
        return None
    source = rows[r]
    if c < 0 or c >= len(source):
        return None
    return source[c]


def _parse_sheet(ws) -> List[Dict[str, Any]]:
    raw_rows = list(ws.iter_rows(values_only=True))
    rows: List[Dict[str, Any]] = []
    row = 1
    while row <= len(raw_rows) - 4:
        arena_id = _to_int(_cell(raw_rows, row, 2), 0)
        if arena_id <= 0:
            row += 1
            continue
        # The sheet stores character name/skill title on the row immediately
        # before the 4-row numeric block.
        name = _text(_cell(raw_rows, row - 1, 4))
        weak_text = _text(_cell(raw_rows, row + 3, 4))
        resist_text = _text(_cell(raw_rows, row + 3, 6))
        skill_text = _text(_cell(raw_rows, row, 8))
        boost_text = _text(_cell(raw_rows, row + 3, 8))
        skill_rate = _to_float(_cell(raw_rows, row + 1, 7), 0.0)
        skill_effects = parse_effect_text(skill_text)
        if skill_rate > 0:
            skill_effects = [effect + [skill_rate] for effect in skill_effects]
        rows.append(
            {
                "sheet": ws.title,
                "character_id": _parse_base_character_id(name),
                "name": name,
                "yang_atk": _to_int(_cell(raw_rows, row, 4), 0),
                "yang_def": _to_int(_cell(raw_rows, row, 6), 0),
                "yin_atk": _to_int(_cell(raw_rows, row + 1, 4), 0),
                "yin_def": _to_int(_cell(raw_rows, row + 1, 6), 0),
                "barrier_count": _to_int(_cell(raw_rows, row + 2, 4), 0),
                "speed": _to_int(_cell(raw_rows, row + 2, 6), 0),
                "weak_text": weak_text,
                "resist_text": resist_text,
                "quality": _quality_from_text(weak_text, resist_text),
                "skill_rate": skill_rate,
                "skill_text": skill_text,
                "skill_effects": skill_effects,
                "boost_text": boost_text,
                "boost_effects": parse_effect_text(boost_text),
                "skill_name": _text(_cell(raw_rows, row - 1, 8)),
            }
        )
        row += 5
    return rows


def parse_arena_workbook(source: Path) -> Dict[str, Any]:
    wb = openpyxl.load_workbook(source, read_only=True, data_only=True)
    rows: List[Dict[str, Any]] = []
    for ws in wb.worksheets[1:3]:
        rows.extend(_parse_sheet(ws))
    multipliers: Dict[str, float] = {}
    if len(wb.worksheets) >= 4:
        ws = wb.worksheets[3]
        for row in ws.iter_rows(values_only=True):
            if row and row[0] is not None:
                multipliers[str(row[0])] = _to_float(row[1], 0.0) if len(row) > 1 else 0.0
    return {
        "parser_version": PARSER_VERSION,
        "source": str(source),
        "sheets": SHEET_ALIASES,
        "multipliers": multipliers,
        "fixed_lunatic_hp": FIXED_LUNATIC_HP,
        "rows": rows,
    }


def write_arena_outputs(payload: Dict[str, Any], csv_path: Path, json_path: Path) -> None:
    json_path.parent.mkdir(parents=True, exist_ok=True)
    csv_path.parent.mkdir(parents=True, exist_ok=True)
    json_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    fieldnames = [
        "sheet",
        "character_id",
        "name",
        "yang_atk",
        "yang_def",
        "yin_atk",
        "yin_def",
        "speed",
        "barrier_count",
        "weak_text",
        "resist_text",
        "skill_rate",
        "skill_text",
        "skill_effects",
        "boost_text",
        "boost_effects",
    ]
    with csv_path.open("w", newline="", encoding="utf-8-sig") as fh:
        writer = csv.DictWriter(fh, fieldnames=fieldnames)
        writer.writeheader()
        for row in payload.get("rows", []):
            flat = {key: row.get(key, "") for key in fieldnames}
            flat["skill_effects"] = json.dumps(flat["skill_effects"], ensure_ascii=False)
            flat["boost_effects"] = json.dumps(flat["boost_effects"], ensure_ascii=False)
            writer.writerow(flat)


def ensure_arena_enemy_data(source: Path, csv_path: Path, json_path: Path) -> Dict[str, Any]:
    if json_path.exists():
        try:
            payload = json.loads(json_path.read_text(encoding="utf-8"))
            if payload.get("parser_version") == PARSER_VERSION:
                return payload
        except Exception:
            pass
    payload = parse_arena_workbook(source)
    write_arena_outputs(payload, csv_path, json_path)
    return payload
