from __future__ import annotations

import csv
import json
import os
from dataclasses import asdict
from typing import Any, Dict, List, Optional, Tuple

from attack_order import get_ally_front_attack_order, parse_attack_segment_order
from battle_op_state import BattleOpStateManager
from buff_effect import parse_buff
from character import Barrier, Buff, CharacterInstanceManager
from combat_constants import BULLET_RAW_LABELS, ELEMENT_RAW_LABELS, TYPE_LABELS
from damage_pipeline_skeleton import execute_attack_phase_for_attacker
from equipment_parser import EQUIPMENT_STAT_CODE_MAP, ensure_equipment_data_json
from gui.config import AllySlotConfig, EnemySlotConfig, ProcessConfig, EQUIPMENT_SLOT_KEYS
from gui.resources import (
    BUFF_TRANSLATION_CSV_PATH,
    BUFF_XLSX_PATH,
    CHARACTER_CSV_PATH,
    DATA_DIR,
    EQUIPMENT_JSON_PATH,
    LOCAL_TRANSLATIONS_JSON_PATH,
    RECOMMENDED_EQUIPMENT_CSV_PATH,
    ROOT_DIR,
    TRANSLATED_CHARACTER_CSV_PATH,
    TRIBE_CSV_PATH,
)


class CharacterIndex:
    def __init__(self, csv_path: str = CHARACTER_CSV_PATH):
        self.csv_path = csv_path
        self.rows: List[Dict[str, Any]] = []
        self.world_to_rows: Dict[str, List[Dict[str, Any]]] = {}
        self._loaded = False

    def _pick_key(self, row: Dict[str, Any], candidates: List[str]) -> Optional[str]:
        lowered = {str(k).strip().lower(): k for k in row.keys()}
        for c in candidates:
            if c.lower() in lowered:
                return lowered[c.lower()]
        return None

    def load(self):
        self.rows = []
        self.world_to_rows = {}
        self._loaded = False
        if not os.path.exists(self.csv_path):
            return
        with open(self.csv_path, "r", encoding="utf-8-sig", newline="") as f:
            reader = csv.DictReader(f)
            for raw in reader:
                if not raw:
                    continue
                id_key = self._pick_key(raw, ["id", "character_id", "角色id", "编号"])
                wg_key = self._pick_key(raw, ["world_group", "世界群组", "world"])
                name_key = self._pick_key(raw, ["name", "角色名", "名字"])
                short_key = self._pick_key(raw, ["short_name", "简称", "短名"])
                if not id_key or not wg_key or not name_key:
                    continue
                try:
                    cid = int(str(raw.get(id_key, "")).strip())
                except Exception:
                    continue
                row = {
                    "id": cid,
                    "world_group": str(raw.get(wg_key, "") or "").strip(),
                    "name": str(raw.get(name_key, "") or "").strip(),
                    "short_name": str(raw.get(short_key, "") or "").strip() if short_key else "",
                }
                self.rows.append(row)
                self.world_to_rows.setdefault(row["world_group"], []).append(row)
        for wg in self.world_to_rows:
            self.world_to_rows[wg].sort(key=lambda x: (x["name"], x["id"]))
        self.rows.sort(key=lambda x: (x["world_group"], x["name"], x["id"]))
        self._loaded = True

    def get_world_groups(self) -> List[str]:
        if not self._loaded:
            self.load()
        return sorted(self.world_to_rows.keys())

    def get_rows_by_world(self, world_group: str) -> List[Dict[str, Any]]:
        if not self._loaded:
            self.load()
        return list(self.world_to_rows.get(world_group, []))


class DamageCalculatorService:
    GUI_HIDDEN_BUFF_IDS = set(range(20, 41)) | {3, 4, 5, 6, 7, 8, 9, 10, 17, 18, 19, 45}
    EQUIPMENT_STYLE_LABELS = {0: "D", 1: "梅", 2: "兰", 3: "菊", 4: "竹"}
    ATTACK_TYPE_TO_EQUIPMENT_SLOT = {
        "1": "1a",
        "2": "2a",
        "5": "5",
    }
    EQUIPMENT_TARGET_LABELS = {
        1: "自身",
        2: "己方全体",
        3: "敌方单体",
        4: "敌方全体",
    }
    STAT_KEY_LABELS = {
        "hp": "HP",
        "yang_atk": "阳攻",
        "yang_def": "阳防",
        "yin_atk": "阴攻",
        "yin_def": "阴防",
        "spd": "速度",
    }
    ATTACK_TYPE_LABELS = {
        "1": "1符",
        "1c": "扩散",
        "2": "2符",
        "2c": "集中",
        "5": "终符",
    }
    CHARACTER_NAME_OVERRIDES = {
        1107: "爱塔妮缇拉尔瓦",
        10107: "爱塔妮缇拉尔瓦",
    }
    COMPOUND_SUB_BUFF_IDS = {1, 2, 41, 42}
    COMPOUND_SUB_VALID_IDS = {1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12}
    ATTACK_BREAK_LABELS = {
        12: "焚灭",
        13: "融冰",
        14: "放电",
        15: "猛毒",
        16: "闪光",
    }
    BIG_BARRIER_LABELS = {
        703701: "L0g大结界: 我方阳攻上升50%，敌方阳攻下降50%",
        703801: "阳防大结界: 我方阳防上升30%，敌方阳防下降30%",
        703901: "全体威力大结界: 我方全体威力上升60%，敌方全体威力下降60%，破界，全体灵力上升2.00",
    }

    def __init__(self, data_dir: str = DATA_DIR):
        self.data_dir = data_dir
        self._tribe_map: Optional[Dict[int, str]] = None
        self._buff_name_map: Optional[Dict[int, str]] = None
        self._buff_sub_name_map: Optional[Dict[Tuple[int, int], str]] = None
        self._buff_target_map: Optional[Dict[int, str]] = None
        self._name_cn_map: Optional[Dict[str, str]] = None
        self._name_cn_by_id: Optional[Dict[int, str]] = None
        self._equipment_by_id: Optional[Dict[int, Dict[str, Any]]] = None
        self._equipment_name_map: Optional[Dict[str, Dict[str, Any]]] = None
        self._recommended_equipment_map: Optional[Dict[int, Dict[str, int]]] = None
        self._local_translation_map: Optional[Dict[str, Dict[str, str]]] = None

    def _load_json_by_id(self, char_id: int) -> Dict[str, Any]:
        path = os.path.join(self.data_dir, f"{char_id}.json")
        if not os.path.exists(path):
            raise FileNotFoundError(f"未找到角色文件: {path}")
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)

    def _resolve_tribe_csv_path(self) -> Optional[str]:
        candidates = [TRIBE_CSV_PATH, os.path.join(ROOT_DIR, "tribe_extracted.csv")]
        for path in candidates:
            if path and os.path.exists(path):
                return path
        return None

    def _ensure_tribe_map(self):
        if self._tribe_map is not None:
            return
        mapping: Dict[int, str] = {}
        tribe_path = self._resolve_tribe_csv_path()
        if tribe_path:
            with open(tribe_path, "r", encoding="utf-8-sig", newline="") as f:
                reader = csv.DictReader(f)
                for row in reader:
                    try:
                        tribe_id = int(str(row.get("ID", "")).strip())
                    except Exception:
                        continue
                    name = str(row.get("tribe_name", "") or "").strip()
                    if name:
                        mapping[tribe_id] = name
        self._tribe_map = mapping

    def _ensure_buff_translation_maps(self):
        if self._buff_name_map is not None and self._buff_sub_name_map is not None and self._buff_target_map is not None:
            return

        buff_name_map: Dict[int, str] = {}
        buff_sub_name_map: Dict[Tuple[int, int], str] = {}
        target_map: Dict[int, str] = {0: "场地", **self.EQUIPMENT_TARGET_LABELS}

        if os.path.exists(BUFF_TRANSLATION_CSV_PATH):
            with open(BUFF_TRANSLATION_CSV_PATH, "r", encoding="utf-8-sig", newline="") as f:
                reader = csv.DictReader(f)
                for row in reader:
                    kind = str(row.get("kind", "") or "").strip().lower()
                    label = str(row.get("label", "") or "").strip()
                    if not label:
                        continue
                    if kind == "target":
                        try:
                            target_map[self._to_int(row.get("sub_id", 0), 0)] = label
                        except Exception:
                            continue
                        continue
                    buff_id = self._to_int(row.get("buff_id", 0), 0)
                    if buff_id <= 0:
                        continue
                    if kind == "buff":
                        buff_name_map[buff_id] = label
                    elif kind == "sub":
                        sub_id = self._to_int(row.get("sub_id", 0), 0)
                        buff_sub_name_map[(buff_id, sub_id)] = label

        if not buff_name_map and os.path.exists(BUFF_XLSX_PATH):
            import openpyxl

            wb = openpyxl.load_workbook(BUFF_XLSX_PATH, data_only=True, read_only=True)
            try:
                ws = wb[wb.sheetnames[1] if len(wb.sheetnames) > 1 else 0]
                current_buff_id: Optional[int] = None
                for row in ws.iter_rows(min_row=2, values_only=True):
                    buff_id = row[2]
                    buff_desc = row[3]
                    sub_id = row[5]
                    sub_desc = row[6]
                    if buff_id is not None:
                        try:
                            current_buff_id = int(buff_id)
                            desc_text = str(buff_desc or "").strip()
                            if desc_text:
                                buff_name_map[current_buff_id] = desc_text
                        except Exception:
                            current_buff_id = None
                    if current_buff_id is None:
                        continue
                    try:
                        sub_id_int = int(sub_id)
                    except Exception:
                        continue
                    sub_desc_text = str(sub_desc or "").strip()
                    if sub_desc_text:
                        buff_sub_name_map[(current_buff_id, sub_id_int)] = sub_desc_text
            finally:
                wb.close()

        self._buff_name_map = buff_name_map
        self._buff_sub_name_map = buff_sub_name_map
        self._buff_target_map = target_map

    def _resolve_name_translation_csv_path(self) -> Optional[str]:
        candidates = [
            TRANSLATED_CHARACTER_CSV_PATH,
            os.path.join(ROOT_DIR, "touhou_characters_translated.csv"),
        ]
        for path in candidates:
            if path and os.path.exists(path):
                return path
        return None

    def _ensure_name_cn_map(self):
        if self._name_cn_map is not None:
            return
        mapping: Dict[str, str] = {}
        id_mapping: Dict[int, str] = {}
        csv_path = self._resolve_name_translation_csv_path()
        if csv_path:
            with open(csv_path, "r", encoding="utf-8-sig", newline="") as f:
                reader = csv.DictReader(f)
                for row in reader:
                    try:
                        char_id = int(str(row.get("id", "") or "").strip())
                    except Exception:
                        char_id = None
                    name = str(row.get("name", "") or "").strip()
                    name_cn = str(row.get("name_cn", "") or "").strip()
                    if name and name_cn:
                        mapping[name] = name_cn
                    if char_id is not None and name_cn:
                        id_mapping[char_id] = name_cn
        for csv_path in [CHARACTER_CSV_PATH, os.path.join(ROOT_DIR, "presets", "characters.csv")]:
            if not csv_path or not os.path.exists(csv_path):
                continue
            with open(csv_path, "r", encoding="utf-8-sig", newline="") as f:
                reader = csv.DictReader(f)
                for row in reader:
                    try:
                        char_id = int(str(row.get("id", "") or "").strip())
                    except Exception:
                        continue
                    name_jp = str(row.get("name_jp", "") or "").strip()
                    name_cn = str(row.get("name", "") or "").strip()
                    if name_jp and name_cn:
                        mapping[name_jp] = name_cn
                    if name_cn:
                        id_mapping[char_id] = name_cn
        for char_id, name_cn in self.CHARACTER_NAME_OVERRIDES.items():
            id_mapping[int(char_id)] = str(name_cn)
        self._name_cn_map = mapping
        self._name_cn_by_id = id_mapping

    def _ensure_local_translation_map(self):
        if self._local_translation_map is not None:
            return
        mapping: Dict[str, Dict[str, str]] = {}
        if os.path.exists(LOCAL_TRANSLATIONS_JSON_PATH):
            with open(LOCAL_TRANSLATIONS_JSON_PATH, "r", encoding="utf-8") as f:
                payload = json.load(f)
            if isinstance(payload, dict):
                for key, value in payload.items():
                    if isinstance(value, dict):
                        mapping[str(key)] = {str(k): str(v) for k, v in value.items()}
        self._local_translation_map = mapping

    def translate_local_text(self, category: str, text: Any) -> str:
        self._ensure_local_translation_map()
        raw = str(text or "").strip()
        if not raw:
            return ""
        return ((self._local_translation_map or {}).get(category, {}) or {}).get(raw, raw)

    def _ensure_equipment_map(self):
        if self._equipment_by_id is not None and self._equipment_name_map is not None:
            return

        ensure_equipment_data_json()
        by_id: Dict[int, Dict[str, Any]] = {}
        by_name: Dict[str, Dict[str, Any]] = {}
        if os.path.exists(EQUIPMENT_JSON_PATH):
            with open(EQUIPMENT_JSON_PATH, "r", encoding="utf-8") as f:
                payload = json.load(f)
            for item in payload.get("items", []) or []:
                try:
                    equipment_id = int(item.get("equipment_id", 0))
                except Exception:
                    continue
                if equipment_id <= 0:
                    continue
                item = dict(item)
                by_id[equipment_id] = item
                name = str(item.get("name", "") or "").strip()
                display_id = str(item.get("equipment_id_text", "") or equipment_id).strip()
                source_id = str(item.get("equipment_id_source_text", "") or "").strip()
                if name:
                    by_name.setdefault(name, item)
                    by_name[f"[{display_id}] {name}"] = item
                    by_name[f"{name} [{display_id}]"] = item
                if display_id:
                    by_name[display_id] = item
                if source_id:
                    by_name[source_id] = item
        self._equipment_by_id = by_id
        self._equipment_name_map = by_name

    def _ensure_recommended_equipment_map(self):
        if self._recommended_equipment_map is not None:
            return
        mapping: Dict[int, Dict[str, int]] = {}
        if os.path.exists(RECOMMENDED_EQUIPMENT_CSV_PATH):
            with open(RECOMMENDED_EQUIPMENT_CSV_PATH, "r", encoding="utf-8-sig", newline="") as f:
                reader = csv.DictReader(f)
                for row in reader:
                    try:
                        char_id = int(str(row.get("character_id", "") or "").strip())
                    except Exception:
                        continue
                    slots: Dict[str, int] = {}
                    for key in EQUIPMENT_SLOT_KEYS:
                        try:
                            slots[key] = int(str(row.get(f"equipment_{key}_id", "") or "0").strip())
                        except Exception:
                            slots[key] = 0
                    if not any(slots.values()):
                        legacy_pairs = {
                            "1a": "1",
                            "2a": "2",
                            "1b": "1c",
                            "2b": "2c",
                            "5": "5",
                        }
                        for key, legacy_key in legacy_pairs.items():
                            try:
                                slots[key] = int(str(row.get(f"equipment_{legacy_key}_id", "") or "0").strip())
                            except Exception:
                                slots[key] = 0
                    mapping[char_id] = slots
        self._recommended_equipment_map = mapping

    def discover_all_ids(self) -> List[int]:
        ids: List[int] = []
        if not os.path.exists(self.data_dir):
            return ids
        for name in os.listdir(self.data_dir):
            if name.endswith(".json") and name[:-5].isdigit():
                ids.append(int(name[:-5]))
        return sorted(ids)

    @staticmethod
    def format_type_label(type_value: Any) -> str:
        try:
            t = int(type_value)
        except Exception:
            return str(type_value or "")
        return TYPE_LABELS.get(t, str(t))

    def load_character_meta(self, char_id: int) -> Dict[str, Any]:
        data = self._load_json_by_id(char_id)
        info = data.get("character_info", {}) or {}
        stats = (data.get("stats", {}) or {}) or info
        attack_skills = data.get("attack_skills", {}) or {}
        raw_type = info.get("type", "")
        return {
            "id": info.get("id", char_id),
            "name": info.get("name", ""),
            "world_group": info.get("world_group", ""),
            "type": raw_type,
            "type_label": self.format_type_label(raw_type),
            "re": info.get("re", False),
            "hp": stats.get("hp", ""),
            "yang_atk": stats.get("yang_atk", ""),
            "yin_atk": stats.get("yin_atk", ""),
            "yang_def": stats.get("yang_def", ""),
            "yin_def": stats.get("yin_def", ""),
            "speed": stats.get("speed", ""),
            "tribe": info.get("tribe", []),
            "available_attack_types": list(attack_skills.keys()),
        }

    @staticmethod
    def _bool_text(value: Any) -> str:
        text = str(value).strip().lower()
        return "是" if text in {"1", "true", "yes"} else "否"

    @staticmethod
    def _to_bool(value: Any) -> bool:
        return str(value).strip().lower() in {"1", "true", "yes"}

    @staticmethod
    def _to_int(value: Any, default: int = 0) -> int:
        try:
            return int(float(value))
        except Exception:
            return default

    def get_character_file_path(self, char_id: int) -> str:
        return os.path.join(self.data_dir, f"{int(char_id)}.json")

    def load_character_full(self, char_id: int) -> Dict[str, Any]:
        return self._load_json_by_id(int(char_id))

    def save_character_full(self, char_id: int, payload: Dict[str, Any]):
        path = self.get_character_file_path(int(char_id))
        with open(path, "w", encoding="utf-8") as f:
            json.dump(payload, f, ensure_ascii=False, indent=4)

    def resolve_character_avatar_path(self, char_id: int) -> Optional[str]:
        candidates = [
            os.path.join(ROOT_DIR, "avatars", f"{int(char_id)}.png"),
            os.path.join(self.data_dir, f"{int(char_id)}.png"),
            os.path.join(ROOT_DIR, f"{int(char_id)}.png"),
        ]
        for path in candidates:
            if os.path.exists(path):
                return path
        return None

    def get_tribe_options(self) -> List[Tuple[int, str]]:
        self._ensure_tribe_map()
        return sorted((self._tribe_map or {}).items(), key=lambda item: item[0])

    def parse_tribe_filter_text(self, text: Any) -> List[int]:
        self._ensure_tribe_map()
        tribe_map = self._tribe_map or {}
        reverse_map = {str(name).strip(): int(tribe_id) for tribe_id, name in tribe_map.items()}
        out: List[int] = []
        seen = set()
        for part in str(text or "").replace("，", ",").replace("；", ",").split(","):
            token = part.strip()
            if not token:
                continue
            tribe_id = None
            if token.endswith(")") and "(" in token:
                maybe_id = token.rsplit("(", 1)[-1].rstrip(")").strip()
                if maybe_id.isdigit():
                    tribe_id = int(maybe_id)
            if tribe_id is None and token.isdigit():
                tribe_id = int(token)
            if tribe_id is None:
                tribe_id = reverse_map.get(token)
            if tribe_id is None or tribe_id in seen:
                continue
            seen.add(tribe_id)
            out.append(int(tribe_id))
        return out

    def get_world_group_options(self) -> List[str]:
        groups = {str(self.load_character_full(char_id).get("character_info", {}).get("world_group", "") or "").strip() for char_id in self.discover_all_ids()}
        return sorted(group for group in groups if group)

    def parse_world_group_filter_text(self, text: Any) -> List[str]:
        available = {group.lower(): group for group in self.get_world_group_options()}
        out: List[str] = []
        seen = set()
        for part in str(text or "").replace("，", ",").replace("；", ",").split(","):
            token = part.strip()
            if not token:
                continue
            group = available.get(token.lower(), token)
            if group in seen:
                continue
            seen.add(group)
            out.append(group)
        return out

    def _get_buff_name(self, buff_id: Any) -> str:
        self._ensure_buff_translation_maps()
        return str((self._buff_name_map or {}).get(self._to_int(buff_id, 0), "") or "")

    def _get_buff_sub_label(self, buff_id: Any, sub_id: Any) -> str:
        self._ensure_buff_translation_maps()
        buff_id_int = self._to_int(buff_id, 0)
        sub_id_int = self._to_int(sub_id, 0)
        if sub_id_int == 13:
            if buff_id_int == 1:
                return "单体减伤/群体易伤"
            if buff_id_int == 2:
                return "群体减伤/单体易伤"
        if buff_id_int == 54 and sub_id_int in self.BIG_BARRIER_LABELS:
            return self.BIG_BARRIER_LABELS[sub_id_int]
        mapping = self._buff_sub_name_map or {}
        label = str(mapping.get((buff_id_int, sub_id_int), "") or "")
        if label:
            return label
        if 21 <= buff_id_int <= 28 or 31 <= buff_id_int <= 38 or buff_id_int in (41, 42, 56, 57):
            label = str(mapping.get((1, sub_id_int), "") or mapping.get((2, sub_id_int), "") or "")
        return label

    def list_buff_id_options(self) -> List[Tuple[int, str]]:
        self._ensure_buff_translation_maps()
        return sorted(
            [
                (int(buff_id), label)
                for buff_id, label in (self._buff_name_map or {}).items()
                if str(label or "").strip() and int(buff_id) not in self.GUI_HIDDEN_BUFF_IDS
            ],
            key=lambda item: item[0],
        )

    def list_buff_subid_options(self, buff_id: Any) -> List[Tuple[int, str]]:
        target_id = self._to_int(buff_id, 0)
        if target_id == 48:
            return [(sub_id, self._get_buff_sub_label(target_id, sub_id) or str(sub_id)) for sub_id in (3, 4, 5)]
        if target_id == 54:
            return [(sub_id, label) for sub_id, label in sorted(self.BIG_BARRIER_LABELS.items())]
        self._ensure_buff_translation_maps()
        options: Dict[int, str] = {}
        for (current_id, sub_id), label in (self._buff_sub_name_map or {}).items():
            if int(current_id) == target_id:
                options[int(sub_id)] = str(label or "")
        if not options and (21 <= target_id <= 28 or 31 <= target_id <= 38 or target_id in (41, 42, 56, 57)):
            for (current_id, sub_id), label in (self._buff_sub_name_map or {}).items():
                if int(current_id) in (1, 2):
                    options.setdefault(int(sub_id), str(label or ""))
        return sorted(options.items(), key=lambda item: item[0])

    def list_equipment_buff_id_options(self) -> List[Tuple[int, str]]:
        self._ensure_equipment_map()
        ids = sorted(
            {
                self._to_int(effect.get("buff_id", 0), 0)
                for item in (self._equipment_by_id or {}).values()
                for effect in (item.get("effects", []) or [])
                if self._to_int(effect.get("buff_id", 0), 0) > 0
            }
        )
        return [(buff_id, self._get_buff_name(buff_id) or str(buff_id)) for buff_id in ids]

    def list_equipment_buff_subid_options(self, buff_id: Any) -> List[Tuple[int, str]]:
        self._ensure_equipment_map()
        target_id = self._to_int(buff_id, 0)
        if target_id == 54:
            return [(sub_id, label) for sub_id, label in sorted(self.BIG_BARRIER_LABELS.items())]
        sub_ids = sorted(
            {
                self._to_int(effect.get("sub_id", 0), 0)
                for item in (self._equipment_by_id or {}).values()
                for effect in (item.get("effects", []) or [])
                if self._to_int(effect.get("buff_id", 0), 0) == target_id
            }
        )
        return [(sub_id, self._get_buff_sub_label(target_id, sub_id) or str(sub_id)) for sub_id in sub_ids]

    def _extract_attack_filter_info(self, attack_skill: Dict[str, Any]) -> Dict[str, Any]:
        attacks = list(attack_skill.get("attacks", []) or [])
        elements = sorted({self._to_int(row.get("element", 0)) for row in attacks if self._to_int(row.get("element", 0)) > 0})
        bullet_types = sorted({self._to_int(row.get("type", 0)) for row in attacks if self._to_int(row.get("type", 0)) > 0})
        killers = sorted({self._to_int(killer) for row in attacks for killer in (row.get("killers", []) or []) if self._to_int(killer) > 0})
        return {
            "elements": elements,
            "bullet_types": bullet_types,
            "killers": killers,
        }

    def _format_attack_element_sequence(self, attack_skill: Dict[str, Any]) -> str:
        attacks = list(attack_skill.get("attacks", []) or [])
        labels = [str(ELEMENT_RAW_LABELS.get(self._to_int(row.get("element", 0)), "")) for row in attacks]
        labels = [label for label in labels if label]
        return "".join(labels) or "-"

    def _format_attack_summary_row(self, attack_type: str, attack_skill: Dict[str, Any]) -> Dict[str, Any]:
        attrs = attack_skill.get("global_attributes", {}) or {}
        info = self._extract_attack_filter_info(attack_skill)
        tribe_map = self._tribe_map or {}
        attack_name = self.translate_local_text("attack_name_by_text", attrs.get("name", ""))
        return {
            "attack_type": str(attack_type),
            "attack_type_label": self.ATTACK_TYPE_LABELS.get(str(attack_type), str(attack_type)),
            "name": str(attack_name or attrs.get("name", "") or ""),
            "attack_count": int(attack_skill.get("attack_count", 0) or 0),
            "elements": " / ".join(ELEMENT_RAW_LABELS.get(v, str(v)) for v in info["elements"]) or "-",
            "bullet_types": " / ".join(BULLET_RAW_LABELS.get(v, str(v)) for v in info["bullet_types"]) or "-",
            "killers": " / ".join(tribe_map.get(v, str(v)) for v in info["killers"]) or "-",
        }

    @staticmethod
    def _format_number(value: Any) -> str:
        try:
            num = float(value)
        except Exception:
            return str(value or "")
        return str(int(num)) if num.is_integer() else f"{num:g}"

    def _format_turn_text(self, duration: Any) -> str:
        duration_int = self._to_int(duration, 0)
        return f"{duration_int}T" if duration_int > 0 else "-"

    def _format_percent_text(self, value: Any) -> str:
        return f"{self._format_number(value)}%"

    def _format_target_prefix(self, target_type: Any) -> str:
        return self.format_equipment_target_label(target_type)

    @staticmethod
    def _format_source_target_stat(sub_label: str) -> Tuple[str, str]:
        text = str(sub_label or "").strip()
        if "→" in text:
            src, dst = text.split("→", 1)
            return src.strip(), dst.strip()
        return text, text

    def _expand_compound_sub_ids(self, buff_id: Any, sub_id: Any) -> List[int]:
        buff_id_int = self._to_int(buff_id, 0)
        sub_id_int = self._to_int(sub_id, 0)
        if buff_id_int not in self.COMPOUND_SUB_BUFF_IDS or sub_id_int <= 100:
            return [sub_id_int]

        text = str(abs(sub_id_int))
        values: List[int] = []
        idx = 0
        while idx < len(text):
            if text[idx] == "0":
                idx += 1
                continue
            picked: Optional[int] = None
            if idx + 2 <= len(text):
                candidate = int(text[idx : idx + 2])
                if candidate in self.COMPOUND_SUB_VALID_IDS:
                    picked = candidate
                    idx += 2
            if picked is None:
                candidate = int(text[idx])
                if candidate in self.COMPOUND_SUB_VALID_IDS:
                    picked = candidate
                idx += 1
            if picked is not None and picked not in values:
                values.append(picked)
        return values or [sub_id_int]

    def _format_multi_sub_label(self, buff_id: Any, sub_id: Any) -> str:
        sub_ids = self._expand_compound_sub_ids(buff_id, sub_id)
        labels = [self._get_buff_sub_label(buff_id, current) or str(current) for current in sub_ids]
        return "、".join(label for label in labels if label) or str(self._to_int(sub_id, 0))

    def _format_effect_text(
        self,
        buff_id: Any,
        sub_id: Any,
        target_type: Any,
        duration: Any,
        value: Any,
        *,
        context: str,
    ) -> str:
        buff_id_int = self._to_int(buff_id, 0)
        sub_id_int = self._to_int(sub_id, 0)
        target_label = self._format_target_prefix(target_type)
        sub_label = self._format_multi_sub_label(buff_id_int, sub_id_int)
        value_text = self._format_number(value)
        percent_text = self._format_percent_text(value)
        turn_text = self._format_turn_text(duration)

        if buff_id_int == 1:
            return f"{target_label}{sub_label}上升{value_text}层 ({turn_text})"
        if buff_id_int == 2:
            return f"{target_label}{sub_label}下降{value_text}层 ({turn_text})"
        if buff_id_int == 3:
            return f"{target_label}体力恢复{percent_text}"
        if buff_id_int == 4:
            return f"{target_label}结界增加{value_text}枚"
        if buff_id_int == 5:
            spirit_value = float(value or 0) / 20.0
            return f"{target_label}灵力上升{self._format_number(spirit_value)}"
        if buff_id_int == 6:
            return f"{target_label}附加{sub_label}{value_text}层"
        if buff_id_int == 7:
            return f"使{target_label}{sub_label}" if context != "equipment" else f"{self._get_buff_name(buff_id_int) or '瞬间效果'} / {sub_label}"
        if buff_id_int == 8:
            return f"使{target_label}{sub_label}" if context != "equipment" else f"{self._get_buff_name(buff_id_int) or '改变行动顺序'} / {sub_label}"
        if buff_id_int == 9:
            return f"使{target_label}解除{value_text}异常结界" if context != "equipment" else f"{self._get_buff_name(buff_id_int) or '解除结界异常'}"
        if buff_id_int == 10:
            return f"使{target_label}解除{value_text}状态异常" if context != "equipment" else f"{self._get_buff_name(buff_id_int) or '解咒'}"
        if buff_id_int in (12, 13, 14):
            return f"{target_label}受到{sub_label}的伤害下降{percent_text} ({turn_text})"
        if buff_id_int in (15, 16):
            return f"{sub_label}的威力上升{percent_text} ({turn_text})"
        if buff_id_int == 17:
            return f"灵力回收上升{percent_text}"
        if buff_id_int == 20:
            return f"{target_label}的{sub_label}上锁"
        if 21 <= buff_id_int <= 28:
            type_label = self.format_type_label(buff_id_int - 20)
            return f"{type_label}使用时，{target_label}{sub_label}上升{value_text}层 ({turn_text})"
        if 31 <= buff_id_int <= 38:
            type_label = self.format_type_label(buff_id_int - 30)
            return f"{type_label}使用时，{target_label}{sub_label}下降{value_text}层 ({turn_text})"
        if buff_id_int == 41:
            return f"{target_label}2阶{sub_label}上升{value_text}层 ({turn_text})"
        if buff_id_int == 42:
            return f"{target_label}2阶{sub_label}下降{value_text}层 ({turn_text})"
        if buff_id_int == 43:
            return f"为{target_label}添加{sub_label}弱点"
        if buff_id_int == 46:
            return f"为{target_label}添加{sub_label}（{turn_text}）"
        if buff_id_int == 47:
            return f"{target_label}{sub_label}{percent_text}"
        if buff_id_int == 48:
            return f"{target_label}根据{sub_label}{percent_text}"
        if buff_id_int == 49:
            return f"{target_label}{sub_label}增伤{percent_text}"
        if buff_id_int == 50:
            return f"{target_label}{sub_label}减伤{percent_text}"
        if buff_id_int == 51:
            return f"{target_label}{sub_label}缩短{percent_text}"
        if buff_id_int == 52:
            return f"{target_label}{sub_label}延长{percent_text}"
        if buff_id_int == 54:
            return f"展开{self._get_buff_sub_label(buff_id_int, sub_id_int) or sub_id_int}"
        if buff_id_int == 56:
            return f"{target_label}{sub_label}永久上升{percent_text}"
        if buff_id_int == 57:
            return f"{target_label}{sub_label}永久下降{percent_text}"
        if buff_id_int == 58:
            source_stat, target_stat = self._format_source_target_stat(sub_label)
            return f"自身的{source_stat}的{percent_text}添加到{target_label}{target_stat}上"

        desc = self.lookup_buff_description(buff_id_int, sub_id_int)
        if target_label and target_label != "场地":
            return f"{target_label}{desc} 值={value_text} 回合={self._to_int(duration, 0)}"
        return f"{desc} 值={value_text} 回合={self._to_int(duration, 0)}"

    def _format_runtime_effect(self, effect_list: Any) -> str:
        if not isinstance(effect_list, (list, tuple)) or len(effect_list) < 5:
            return "-"
        return self._format_effect_text(
            effect_list[0],
            effect_list[1],
            effect_list[2],
            effect_list[3],
            effect_list[4],
            context="skill",
        )

    def _format_skill_entry(self, skill: Dict[str, Any]) -> Dict[str, str]:
        skill_name = self.translate_local_text("skill_name_by_text", skill.get("name", ""))
        parts = []
        for key in ("a", "b", "c"):
            effect = skill.get(key, [])
            if effect:
                parts.append(f"{key.upper()}. {self._format_runtime_effect(effect)}")
        title = f"技能{self._to_int(skill.get('id', 0), 0) or '-'} | {skill_name or '-'}"
        return {
            "title": title,
            "content": "\n".join(parts) if parts else "无效果说明",
        }

    def _format_passive_entry(self, info: Dict[str, Any], idx: int) -> str:
        name = self.translate_local_text("passive_name_by_text", info.get(f"passivity_{idx}_name", ""))
        desc = str(info.get(f"passivity_{idx}_description", "") or "")
        return f"{idx}. {name} | {desc or '-'}"

    def _format_attack_paragraph_buffs(self, hit: Dict[str, Any]) -> str:
        raw_buffs = hit.get("buff", []) or []
        if not isinstance(raw_buffs, list) or not raw_buffs:
            return "-"
        parts: List[str] = []
        for row in parse_buff(raw_buffs):
            if not row or len(row) < 5:
                continue
            parts.append(
                self._format_effect_text(
                    row[0],
                    row[1],
                    row[2],
                    row[3],
                    row[4],
                    context="skill",
                )
            )
        return " / ".join(part for part in parts if part and part != "-") or "-"

    def _format_attack_effect_details(self, hit: Dict[str, Any]) -> Dict[str, str]:
        traits: List[str] = []
        raw_effects = hit.get("effect", []) or []
        if not isinstance(raw_effects, list):
            return {"traits": "-", "buffs": self._format_attack_paragraph_buffs(hit)}
        for idx in range(0, len(raw_effects), 2):
            if idx + 1 >= len(raw_effects):
                break
            effect_id = self._to_int(raw_effects[idx], 0)
            value = raw_effects[idx + 1]
            value_text = self._format_number(value)
            if effect_id == 1:
                traits.append("必中")
            elif effect_id == 2:
                traits.append("贯穿")
            elif effect_id == 3:
                traits.append("镜面")
            elif effect_id == 4:
                traits.append(f"硬质{value_text}%")
            elif effect_id == 5:
                traits.append(f"斩裂{value_text}%")
            elif effect_id == 7:
                traits.append("爆破")
            elif effect_id == 8:
                traits.append("弹性")
            elif effect_id == 9:
                traits.append("精密")
            elif effect_id == 10:
                traits.append(f"特性10({value_text})")
            elif effect_id == 11:
                traits.append(f"吸血{value_text}%")
            elif 12 <= effect_id <= 16:
                traits.append(self.ATTACK_BREAK_LABELS.get(effect_id, f"击破{effect_id}"))
            elif effect_id > 0:
                traits.append(f"effect {effect_id}({value_text})")
        return {
            "traits": " / ".join(traits) if traits else "-",
            "buffs": self._format_attack_paragraph_buffs(hit),
        }

    def _format_attack_buff_entry(self, attack_type: str, attack_skill: Dict[str, Any]) -> str:
        attrs = attack_skill.get("global_attributes", {}) or {}
        before = []
        after = []
        for row in list(attrs.get("effect_before", []) or []):
            if row:
                before.append(self._format_runtime_effect(row))
        for row in list(attrs.get("effect_after", []) or []):
            if row:
                after.append(self._format_runtime_effect(row))
        label = self.ATTACK_TYPE_LABELS.get(str(attack_type), str(attack_type))
        attack_name = self.translate_local_text("attack_name_by_text", attrs.get("name", ""))
        return f"{label} / {attack_name}\n攻击前: {('；'.join(before) if before else '-')}\n攻击后: {('；'.join(after) if after else '-')}"

    def _format_attack_section(self, attack_type: str, attack_skill: Dict[str, Any]) -> Dict[str, str]:
        attrs = attack_skill.get("global_attributes", {}) or {}
        section_title = self.ATTACK_TYPE_LABELS.get(str(attack_type), str(attack_type))
        attack_name = self.translate_local_text("attack_name_by_text", attrs.get("name", ""))
        attack_info = self._format_attack_summary_row(attack_type, attack_skill)
        before = []
        after = []
        for row in list(attrs.get("effect_before", []) or []):
            if row:
                before.append(self._format_runtime_effect(row))
        for row in list(attrs.get("effect_after", []) or []):
            if row:
                after.append(self._format_runtime_effect(row))

        hit_lines: List[str] = []
        tribe_map = self._tribe_map or {}
        for hit in list(attack_skill.get("attacks", []) or []):
            shot_name = self.translate_local_text("shot_name_by_text", hit.get("name", ""))
            element_label = ELEMENT_RAW_LABELS.get(self._to_int(hit.get("element", 0)), str(hit.get("element", "")))
            bullet_label = BULLET_RAW_LABELS.get(self._to_int(hit.get("type", 0)), str(hit.get("type", "")))
            killer_labels = " / ".join(tribe_map.get(self._to_int(k), str(k)) for k in (hit.get("killers", []) or []) if self._to_int(k) > 0) or "-"
            yy = "阳" if self._to_int(hit.get("yinyang", 0)) == 1 else "阴"
            per_hit_power = float(hit.get("damage", 0) or 0)
            hit_count = self._to_int(hit.get("amt", 0), 0)
            total_power = per_hit_power * hit_count
            order_value = str(hit.get("order", "") or "").strip()
            order_label = order_value if order_value else str(hit.get("attack_id", "") or self._to_int(hit.get("id", 0), 0))
            acc_rate = self._format_number(hit.get("acc", 0))
            cri_rate = self._format_number(hit.get("cri", 0))
            effect_detail = self._format_attack_effect_details(hit)
            hit_lines.append(
                f"段{order_label} {shot_name} | {element_label}|{bullet_label}|{yy} | 特攻={killer_labels}\n"
                f"威力 {self._format_number(per_hit_power)} | Hit {hit_count} | 总威力 {self._format_number(total_power)} | 命中率 {acc_rate}% | 会心率 {cri_rate}% | 特性: {effect_detail['traits']}\n"
                f"效果: {effect_detail['buffs']}"
            )

        detail_text = "\n".join(
            [
                f"名称: {attack_name}",
                f"段数: {attack_info['attack_count']}",
                f"属性: {attack_info['elements']}",
                f"弹种: {attack_info['bullet_types']}",
                f"特攻: {attack_info['killers']}",
                f"攻击前Buff: {('；'.join(before) if before else '-')}",
                f"攻击后Buff: {('；'.join(after) if after else '-')}",
                "Hit详情:",
                *(hit_lines or ["-"]),
            ]
        )
        return {
            "title": section_title,
            "name": attack_name,
            "content": detail_text,
        }

    def get_character_detail_payload(self, char_id: int) -> Dict[str, Any]:
        self._ensure_tribe_map()
        data = self.load_character_full(char_id)
        info = data.get("character_info", {}) or {}
        stats = (data.get("stats", {}) or {}) or info
        attack_skills = data.get("attack_skills", {}) or {}
        translated_name = self.translate_character_name(info.get("name", ""), char_id)
        return {
            "char_id": int(char_id),
            "path": self.get_character_file_path(int(char_id)),
            "avatar_path": self.resolve_character_avatar_path(int(char_id)),
            "name": str(translated_name or info.get("name", "") or ""),
            "name_raw": str(info.get("name", "") or ""),
            "subname": self.translate_local_text("subname_by_text", info.get("subname", "")),
            "world_group": str(info.get("world_group", "") or ""),
            "type": self._to_int(info.get("type", 0)),
            "type_label": self.format_type_label(info.get("type", 0)),
            "re": self._to_bool(info.get("re", False)),
            "hp": self._to_int(stats.get("hp", 0)),
            "yang_atk": self._to_int(stats.get("yang_atk", 0)),
            "yang_def": self._to_int(stats.get("yang_def", 0)),
            "yin_atk": self._to_int(stats.get("yin_atk", 0)),
            "yin_def": self._to_int(stats.get("yin_def", 0)),
            "speed": self._to_int(stats.get("speed", 0)),
            "tribe_ids": list(info.get("tribe", []) or []),
            "tribe_text": self.describe_tribe_text(",".join(str(v) for v in (info.get("tribe", []) or []))),
            "attack_rows": [self._format_attack_summary_row(key, value) for key, value in attack_skills.items()],
            "attack_sections": [self._format_attack_section(key, value) for key, value in attack_skills.items()],
            "skill_entries": [self._format_skill_entry(skill) for skill in (data.get("skills", []) or [])],
            "passive_entries": [self._format_passive_entry(info, idx) for idx in range(1, 4)],
            "raw_json_text": json.dumps(data, ensure_ascii=False, indent=4),
        }

    def search_characters(
        self,
        *,
        type_values: Optional[List[Any]] = None,
        element_values: Optional[List[Any]] = None,
        bullet_values: Optional[List[Any]] = None,
        killer_tribes: Optional[List[Any]] = None,
        world_groups: Optional[List[str]] = None,
        re_only: str = "",
    ) -> List[Dict[str, Any]]:
        results: List[Dict[str, Any]] = []
        type_filters = {self._to_int(value, 0) for value in (type_values or []) if self._to_int(value, 0) > 0}
        element_filters = {self._to_int(value, 0) for value in (element_values or []) if self._to_int(value, 0) > 0}
        bullet_filters = {self._to_int(value, 0) for value in (bullet_values or []) if self._to_int(value, 0) > 0}
        killer_filters = {self._to_int(value, 0) for value in (killer_tribes or []) if self._to_int(value, 0) > 0}
        world_group_filters = {str(value).strip() for value in (world_groups or []) if str(value).strip()}
        re_filter = str(re_only or "").strip()
        for char_id in self.discover_all_ids():
            data = self.load_character_full(char_id)
            info = data.get("character_info", {}) or {}
            stats = (data.get("stats", {}) or {}) or info
            attack5 = (data.get("attack_skills", {}) or {}).get("5", {}) or {}
            attack5_info = self._extract_attack_filter_info(attack5)

            raw_type = self._to_int(info.get("type", 0))
            world_group = str(info.get("world_group", "") or "").strip()
            is_re = self._to_bool(info.get("re", False))
            if type_filters and raw_type not in type_filters:
                continue
            if world_group_filters and world_group not in world_group_filters:
                continue
            if element_filters and not element_filters.intersection(set(attack5_info["elements"])):
                continue
            if bullet_filters and not bullet_filters.intersection(set(attack5_info["bullet_types"])):
                continue
            if killer_filters and not killer_filters.intersection(set(attack5_info["killers"])):
                continue
            if re_filter == "是" and not is_re:
                continue
            if re_filter == "否" and is_re:
                continue

            results.append(
                {
                    "character_id": int(char_id),
                    "name": self.translate_character_name(info.get("name", ""), char_id),
                    "world_group": world_group,
                    "type": raw_type,
                    "type_label": self.format_type_label(raw_type),
                    "yang_atk": self._to_int(stats.get("yang_atk", 0)),
                    "yang_def": self._to_int(stats.get("yang_def", 0)),
                    "yin_atk": self._to_int(stats.get("yin_atk", 0)),
                    "yin_def": self._to_int(stats.get("yin_def", 0)),
                    "speed": self._to_int(stats.get("speed", 0)),
                    "re": is_re,
                    "attack5_element_sequence": self._format_attack_element_sequence(attack5),
                }
            )
        results.sort(key=lambda row: (row["name"], row["character_id"]))
        return results

    def format_equipment_target_label(self, target_type: Any) -> str:
        self._ensure_buff_translation_maps()
        target_type_int = self._to_int(target_type, 0)
        return str((self._buff_target_map or {}).get(target_type_int, self.EQUIPMENT_TARGET_LABELS.get(target_type_int, str(target_type or "-"))) or "-")

    def format_equipment_effect_text(self, effect: Dict[str, Any]) -> str:
        if not effect:
            return "-"
        if bool(effect.get("ignored", False)) or self._to_int(effect.get("raw_id", 0), 0) == 0 or self._to_int(effect.get("buff_id", 0), 0) == 0:
            return "-"
        text = self._format_effect_text(
            effect.get("buff_id", 0),
            effect.get("sub_id", 0),
            effect.get("target_type", 0),
            effect.get("duration", 0),
            effect.get("value", 0),
            context="equipment",
        )
        condition_type = self._to_int(effect.get("condition_type", 0), 0)
        if condition_type > 0 and self._to_int(effect.get("buff_id", 0), 0) in (1, 2):
            return f"{self.format_type_label(condition_type)}使用时，{text}"
        return text

    def search_equipment(
        self,
        *,
        stars: Any = "",
        style_code: Any = "",
        buff_filters: Optional[List[Dict[str, Any]]] = None,
        stat_filters: Optional[List[str]] = None,
    ) -> List[Dict[str, Any]]:
        self._ensure_equipment_map()
        star_filter = self._to_int(stars, 0)
        style_filter = self._to_int(style_code, -1)
        normalized_buff_filters = buff_filters or []
        normalized_stat_filters = [str(key).strip() for key in (stat_filters or []) if str(key).strip()]

        rows: List[Dict[str, Any]] = []
        for item in sorted((self._equipment_by_id or {}).values(), key=lambda row: int(row.get("equipment_id", 0))):
            if star_filter and self._to_int(item.get("stars", 0)) != star_filter:
                continue
            if style_filter >= 0 and self._to_int(item.get("card_style_code", 0)) != style_filter:
                continue

            stats = list(item.get("stats", []) or [])
            stat_keys = {str(row.get("stats_key", "") or "").strip() for row in stats}
            if any(key not in stat_keys for key in normalized_stat_filters):
                continue

            effects = list(item.get("effects", []) or [])
            usable_effects = [effect for effect in effects if not bool(effect.get("ignored", False)) and self._to_int(effect.get("buff_id", 0), 0) > 0]
            matched = True
            for filter_item in normalized_buff_filters:
                buff_id_filter = self._to_int(filter_item.get("buff_id", 0), 0)
                sub_id_filters = [self._to_int(value, 0) for value in (filter_item.get("sub_ids", []) or []) if self._to_int(value, 0) > 0]
                target_filter = self._to_int(filter_item.get("target_type", 0), 0)
                type_filters = [self._to_int(value, 0) for value in (filter_item.get("type_conditions", []) or []) if self._to_int(value, 0) > 0]
                value_filter_text = str(filter_item.get("value", "") or "").strip()
                has_match = any(
                    (not buff_id_filter or self._to_int(effect.get("buff_id", 0), 0) == buff_id_filter)
                    and (not sub_id_filters or self._to_int(effect.get("sub_id", 0), 0) in sub_id_filters)
                    and (not target_filter or self._to_int(effect.get("target_type", 0), 0) == target_filter)
                    and (
                        not type_filters
                        or buff_id_filter not in (1, 2)
                        or self._to_int(effect.get("condition_type", 0), 0) in type_filters
                    )
                    and (
                        not value_filter_text
                        or (
                            self._is_same_numeric_value(effect.get("value", 0), value_filter_text)
                        )
                    )
                    for effect in usable_effects
                )
                if not has_match:
                    matched = False
                    break
            if not matched:
                continue

            rows.append(
                {
                    "equipment_id": self._to_int(item.get("equipment_id", 0)),
                    "name": str(item.get("name", "") or ""),
                    "stars": self._to_int(item.get("stars", 0)),
                    "style_code": self._to_int(item.get("card_style_code", 0)),
                    "style_label": self.EQUIPMENT_STYLE_LABELS.get(self._to_int(item.get("card_style_code", 0)), str(item.get("card_style_label", "") or "")),
                    "stats_text": "\n".join(
                        f"{self.STAT_KEY_LABELS.get(str(row.get('stats_key', '') or '').strip(), row.get('label', ''))}+{self._to_int(row.get('value', 0))}"
                        for row in stats
                    )
                    or "-",
                    "buff_1_text": self.format_equipment_effect_text(effects[0]) if len(effects) > 0 else "-",
                    "buff_2_text": self.format_equipment_effect_text(effects[1]) if len(effects) > 1 else "-",
                    "buff_3_text": self.format_equipment_effect_text(effects[2]) if len(effects) > 2 else "-",
                    "effects_text": " | ".join(self.format_equipment_effect_text(effect) for effect in effects) or "-",
                    "raw_effects": effects[:3],
                }
            )
        rows.sort(key=lambda row: (row["name"], row["equipment_id"]))
        return rows

    @staticmethod
    def _is_same_numeric_value(left: Any, right: Any) -> bool:
        try:
            return abs(float(left) - float(right)) <= 1e-9
        except Exception:
            return False

    def describe_tribe_text(self, text: str) -> str:
        self._ensure_tribe_map()
        tribe_map = self._tribe_map or {}
        labels: List[str] = []
        seen = set()
        for part in str(text or "").replace("，", ",").split(","):
            s = part.strip()
            if not s:
                continue
            try:
                tribe_id = int(s)
            except Exception:
                labels.append(f"{s}=未知")
                continue
            if tribe_id in seen:
                continue
            seen.add(tribe_id)
            labels.append(f"{tribe_id}:{tribe_map.get(tribe_id, '未知')}")
        return " / ".join(labels) if labels else "-"

    def lookup_buff_description(self, buff_id: Any, sub_id: Any) -> str:
        buff_id_int = self._to_int(buff_id, 0)
        sub_id_int = self._to_int(sub_id, 0)
        buff_name = self._get_buff_name(buff_id_int)
        sub_label = self._get_buff_sub_label(buff_id_int, sub_id_int)
        if buff_name and sub_label:
            return f"{buff_name} / {sub_label}"
        if sub_label:
            return sub_label
        if buff_name:
            return buff_name
        return "未知"

    def translate_character_name(self, name: Any, char_id: Any = None) -> str:
        self._ensure_name_cn_map()
        text = str(name or "").strip()
        if char_id not in (None, ""):
            try:
                char_id_int = int(float(char_id))
                translated = (self._name_cn_by_id or {}).get(char_id_int)
            except Exception:
                char_id_int = None
                translated = None
            if char_id_int is not None:
                self._ensure_local_translation_map()
                translated = ((self._local_translation_map or {}).get("character_name_by_id", {}) or {}).get(str(char_id_int), translated)
            if translated:
                return translated
        if not text:
            return ""
        translated = (self._name_cn_map or {}).get(text, text)
        return self.translate_local_text("character_name_by_text", translated)

    def get_equipment(self, equipment_id: Any) -> Optional[Dict[str, Any]]:
        self._ensure_equipment_map()
        raw = str(equipment_id or "").strip()
        if raw and raw in (self._equipment_name_map or {}):
            return (self._equipment_name_map or {}).get(raw)
        try:
            key = int(float(equipment_id))
        except Exception:
            return None
        return (self._equipment_by_id or {}).get(key)

    def resolve_equipment_by_text(self, text: str) -> Optional[Dict[str, Any]]:
        self._ensure_equipment_map()
        raw = str(text or "").strip()
        if not raw:
            return None
        direct = (self._equipment_name_map or {}).get(raw)
        if direct is not None:
            return direct
        try:
            return self.get_equipment(int(raw))
        except Exception:
            pass
        return (self._equipment_name_map or {}).get(raw)

    def list_equipment_options(self) -> List[Tuple[str, int]]:
        self._ensure_equipment_map()
        items = sorted((self._equipment_by_id or {}).values(), key=lambda item: int(item.get("equipment_id", 0)))
        return [
            (f"[{item.get('equipment_id_text', item['equipment_id'])}] {item['name']}", int(item["equipment_id"]))
            for item in items
        ]

    def describe_equipment_text(self, equipment_id: Any) -> str:
        equipment = self.get_equipment(equipment_id)
        if not equipment:
            return "无绘卷"
        stat_parts = [f"{row.get('label', '')}+{int(row.get('value', 0))}" for row in equipment.get("stats", []) or []]
        effect_parts: List[str] = []
        for effect in equipment.get("effects", []) or []:
            if effect.get("ignored"):
                continue
            raw_id = int(effect.get("raw_id", 0) or 0)
            buff_id = int(effect.get("buff_id", 0) or 0)
            if raw_id >= 2000:
                effect_parts.append(f"专属效果 {self.format_equipment_effect_text(effect)}")
            elif buff_id > 0:
                effect_parts.append(self.format_equipment_effect_text(effect))
        parts = []
        if stat_parts:
            parts.append("面板 " + " / ".join(stat_parts))
        if effect_parts:
            parts.append("效果 " + " / ".join(effect_parts))
        return f"{equipment['name']} | " + (" | ".join(parts) if parts else "无额外效果")

    def get_recommended_equipment_id(self, character_id: Any, slot_key: str) -> int:
        self._ensure_recommended_equipment_map()
        try:
            cid = int(float(character_id))
        except Exception:
            return 0
        return int(((self._recommended_equipment_map or {}).get(cid, {}) or {}).get(str(slot_key), 0) or 0)

    def fill_missing_recommended_equipment_ids(self, character_id: Any, equipment_ids: Dict[str, Any]) -> Dict[str, int]:
        merged = {key: 0 for key in EQUIPMENT_SLOT_KEYS}
        raw = equipment_ids or {}
        if isinstance(raw, dict):
            for key in EQUIPMENT_SLOT_KEYS:
                try:
                    merged[key] = int(raw.get(key, 0) or 0)
                except Exception:
                    merged[key] = 0
        try:
            cid = int(float(character_id))
        except Exception:
            return merged
        self._ensure_recommended_equipment_map()
        recommended = (self._recommended_equipment_map or {}).get(cid, {}) or {}
        for key in EQUIPMENT_SLOT_KEYS:
            if merged[key] > 0:
                continue
            try:
                merged[key] = int(recommended.get(key, 0) or 0)
            except Exception:
                merged[key] = 0
        return merged

    @staticmethod
    def get_equipment_slot_key(attack_type: str) -> str:
        attack_type = str(attack_type or "").strip()
        return DamageCalculatorService.ATTACK_TYPE_TO_EQUIPMENT_SLOT.get(attack_type, "")

    def get_equipment_for_attack_type(self, equipment_ids: Dict[str, Any], attack_type: str) -> Optional[Dict[str, Any]]:
        slot_key = self.get_equipment_slot_key(attack_type)
        if not slot_key:
            return None
        equipment_id = int((equipment_ids or {}).get(slot_key, 0) or 0)
        return self.get_equipment(equipment_id)

    @staticmethod
    def parse_skill_order(text: str, front_pos: int) -> List[List[int]]:
        text = (text or "").strip()
        if not text:
            return []
        nums = [int(x.strip()) for x in text.split(",") if x.strip()]
        return [[front_pos, n] for n in nums]

    @staticmethod
    def parse_global_skill_order(text: str, enabled_allies: List[int]) -> List[List[int]]:
        text = (text or "").strip()
        if not text:
            raise ValueError("已启用自定义技能顺序，但顺序文本为空")
        allowed = set(enabled_allies)
        parts = [p.strip() for p in text.replace("\n", ",").replace(";", ",").replace("；", ",").split(",") if p.strip()]
        order: List[List[int]] = []
        for part in parts:
            if ":" not in part:
                raise ValueError(f"技能顺序项格式错误: {part}，应为 位置:技能位")
            pos_txt, skill_txt = part.split(":", 1)
            pos = int(pos_txt.strip())
            skill = int(skill_txt.strip())
            if pos not in allowed:
                raise ValueError(f"技能顺序里引用了未启用角色位置: {pos}")
            if skill < 0:
                raise ValueError(f"技能位不能为负数: {skill}")
            order.append([pos, skill])
        return order

    @staticmethod
    def parse_tribe(text: str) -> List[int]:
        vals: List[int] = []
        for part in str(text or "").replace("，", ",").split(","):
            s = part.strip()
            if not s:
                continue
            vals.append(int(s))
        return vals

    @staticmethod
    def parse_buff_rows(rows: List[List[Any]]) -> List[Buff]:
        out: List[Buff] = []
        for row in rows:
            if not row or len(row) < 4:
                continue
            out.append(Buff(buff_id=int(row[0]), sub_id=int(row[1]), duration=int(row[2]), value=float(row[3])))
        return out

    def _apply_equipment_stats(self, ally, equipment: Dict[str, Any]):
        stats = getattr(ally, "stats", None)
        if stats is None:
            return
        for row in equipment.get("stats", []) or []:
            key = str(row.get("stats_key", "") or "").strip()
            value = float(row.get("value", 0) or 0)
            if not key or not hasattr(stats, key):
                continue
            current = float(getattr(stats, key, 0) or 0)
            setattr(stats, key, int(round(current + value)))

    def _should_apply_equipment_effect(self, ally, equipment: Dict[str, Any], effect: Dict[str, Any], mode: str) -> bool:
        if not effect or effect.get("ignored"):
            return False

        buff_id = int(effect.get("buff_id", 0) or 0)
        raw_id = int(effect.get("raw_id", 0) or 0)
        if mode == "panel_only":
            return False
        if mode == "panel_and_attack_bonus" and buff_id not in (15, 16):
            return False

        info = getattr(ally, "character_info", {}) or {}
        char_type = int(info.get("type", 0) or 0)
        condition_type = int(effect.get("condition_type", 0) or 0)
        if condition_type and char_type != condition_type:
            return False

        exclusive_character_id = int(equipment.get("exclusive_character_id", 0) or 0)
        if raw_id >= 2000:
            if exclusive_character_id <= 0:
                return False
            try:
                current_id = int(info.get("id", 0) or 0)
            except Exception:
                current_id = 0
            if current_id != exclusive_character_id:
                return False

        return True

    def _apply_selected_equipment(
        self,
        mgr: CharacterInstanceManager,
        op: BattleOpStateManager,
        ally_pos: int,
        cfg: AllySlotConfig,
        *,
        mode: str = "full",
    ) -> Dict[str, Any]:
        slot_key = self.get_equipment_slot_key(cfg.attack_type)
        if not slot_key:
            return {"equipment_id": 0, "name": "", "slot_key": ""}
        equipment_id = int((cfg.equipment_ids or {}).get(slot_key, 0) or 0)
        if equipment_id <= 0:
            return {"equipment_id": 0, "name": "", "slot_key": slot_key}

        equipment = self.get_equipment(equipment_id)
        ally = mgr.get_character_instance("我方前排", ally_pos)
        if not equipment or ally is None:
            return {"equipment_id": equipment_id, "name": "", "slot_key": slot_key}

        self._apply_equipment_stats(ally, equipment)
        enemy_target_pos = int(cfg.target_enemy_pos)
        for effect in equipment.get("effects", []) or []:
            if not self._should_apply_equipment_effect(ally, equipment, effect, mode):
                continue
            op._apply_buff_effect(
                effect_list=[
                    int(effect.get("buff_id", 0) or 0),
                    int(effect.get("sub_id", 0) or 0),
                    int(effect.get("target_type", 1) or 1),
                    int(effect.get("duration", 0) or 0),
                    float(effect.get("value", 0) or 0.0),
                ],
                char_manager=mgr,
                attacker_pos_idx=ally_pos,
                enemy_target_pos_idx=enemy_target_pos,
            )

        return {
            "equipment_id": int(equipment["equipment_id"]),
            "name": str(equipment.get("name", "") or ""),
            "slot_key": slot_key,
        }

    def _apply_enemy_slot(self, mgr: CharacterInstanceManager, enemy_pos: int, cfg: EnemySlotConfig):
        if not cfg.enabled:
            return
        mgr.add_enemy_by_id(
            position=enemy_pos,
            character_id=cfg.character_id,
            data_dir=self.data_dir,
            yang_def_override=cfg.yang_def,
            yin_def_override=cfg.yin_def,
            hp_override=cfg.hp,
            barrier_count=cfg.barrier_count,
        )
        enemy = mgr.get_character_instance("敌方", enemy_pos)
        if enemy is None:
            raise RuntimeError(f"敌方位置 {enemy_pos} 加载失败")
        enemy.stats.hp = int(cfg.hp)
        enemy.stats.yang_def = int(cfg.yang_def)
        enemy.stats.yin_def = int(cfg.yin_def)
        enemy.stats.quality = [1] + list(cfg.quality)
        enemy.character_info["tribe"] = self.parse_tribe(cfg.tribe_text)
        enemy.barriers = [Barrier(barrier_id=i + 1, type=0, is_active=True) for i in range(int(cfg.barrier_count))]
        enemy.buffs = self.parse_buff_rows(cfg.buffs)
        enemy.is_break_all = False
        if cfg.is_break_all:
            for b in enemy.barriers:
                b.is_active = False
                b.type = 0
            enemy.break_all_barrier()

    def _apply_ally_slot(self, mgr: CharacterInstanceManager, ally_pos: int, cfg: AllySlotConfig):
        if not cfg.enabled:
            return
        ally_data = self._load_json_by_id(cfg.character_id)
        mgr.add_ally_character(position=ally_pos, char_data=ally_data)
        ally = mgr.get_character_instance("我方前排", ally_pos)
        if ally is None:
            raise RuntimeError(f"我方位置 {ally_pos} 加载失败")
        ally.update_spirit(float(cfg.initial_spirit), "set")
        ally.barriers = [Barrier(barrier_id=i + 1, type=0, is_active=True) for i in range(int(cfg.barrier_count))]
        ally.is_break_all = False
        ally.buffs = list(getattr(ally, "buffs", []) or []) + self.parse_buff_rows(cfg.buffs)

    def run_single(
        self,
        enemy_slots: Dict[int, EnemySlotConfig],
        ally_slots: Dict[int, AllySlotConfig],
        process_config: Optional[ProcessConfig] = None,
        equipment_mode: str = "full",
    ) -> Dict[str, Any]:
        process_config = process_config or ProcessConfig()
        mgr = CharacterInstanceManager()
        op = BattleOpStateManager()

        enabled_enemies = [pos for pos, cfg in enemy_slots.items() if cfg.enabled]
        if not enabled_enemies:
            raise RuntimeError("至少启用一个敌人")
        for enemy_pos in [0, 1, 2]:
            self._apply_enemy_slot(mgr, enemy_pos, enemy_slots[enemy_pos])

        enabled_allies = [pos for pos, cfg in ally_slots.items() if cfg.enabled]
        if not enabled_allies:
            raise RuntimeError("至少启用一个我方前排角色")
        for ally_pos in [0, 1, 2]:
            self._apply_ally_slot(mgr, ally_pos, ally_slots[ally_pos])

        for ally_pos in enabled_allies:
            cfg = ally_slots[ally_pos]
            op.set_enemy_target_pos(ally_pos, int(cfg.target_enemy_pos))
            op.set_shield_open_count(ally_pos, int(cfg.shield_open_count))
            op.set_attack_params(ally_pos, str(cfg.attack_type), int(cfg.spirit_level))

        if process_config.use_custom_skill_order and len(enabled_allies) > 1:
            skill_order = self.parse_global_skill_order(process_config.custom_skill_order_text, enabled_allies)
        else:
            skill_order: List[List[int]] = []
            for ally_pos in enabled_allies:
                cfg = ally_slots[ally_pos]
                skill_order.extend(self.parse_skill_order(cfg.skill_order_text, ally_pos))

        op.set_skill_order(skill_order)
        op.set_field_buffs(process_config.field_buffs.to_runtime_dict())

        ally_front_chars = [mgr.get_character_instance("我方前排", i) for i in (0, 1, 2)]
        enemy_chars = [mgr.get_character_instance("敌方", i) for i in (0, 1, 2)]

        op_result = {
            "success": True,
            "errors": [],
            "skill_execution": [],
            "shield_execution": [],
            "spirit_execution": [],
            "pre_attack_buff_execution": [],
            "post_attack_buff_execution": [],
        }

        op._execute_skills(mgr, ally_front_chars, enemy_chars, op_result)
        op._execute_shields(mgr, ally_front_chars, op_result)
        op._execute_spirit(mgr, ally_front_chars, op_result)

        selected_equipment: Dict[int, Dict[str, Any]] = {}
        for ally_pos in enabled_allies:
            selected_equipment[ally_pos] = self._apply_selected_equipment(
                mgr,
                op,
                ally_pos,
                ally_slots[ally_pos],
                mode=equipment_mode,
            )

        order, snapshot = get_ally_front_attack_order(mgr, return_debug=True)
        all_details: List[Dict[str, Any]] = []
        ally_totals = {0: 0, 1: 0, 2: 0}
        ally_seg_damage = {0: {f"seg{i}": 0 for i in range(6)}, 1: {f"seg{i}": 0 for i in range(6)}, 2: {f"seg{i}": 0 for i in range(6)}}

        for attacker_pos in order:
            attacker = mgr.get_character_instance("我方前排", attacker_pos)
            if attacker is None:
                continue
            op_status = op.get_op_status(attacker_pos)
            attack_type = op_status.attack_type
            spirit_level = int(getattr(op_status, "spirit_level", 0))
            if attack_type not in (attacker.attack_skills or {}):
                continue
            attack_skill_data = attacker.attack_skills[attack_type]
            segment_order = parse_attack_segment_order(attack_skill_data, spirit_level)
            op.execute_pre_attack_buffs_for_attacker(mgr, attacker_pos, op_result)
            details = execute_attack_phase_for_attacker(
                mgr,
                op,
                attacker_pos=attacker_pos,
                segment_order=segment_order,
                debug_print=False,
            )
            for d in details or []:
                all_details.append(d)
                dmg = int(d.get("damage_int", 0) or 0)
                seg = int(d.get("seg", -1))
                ally_totals[attacker_pos] += dmg
                if 0 <= seg < 6:
                    ally_seg_damage[attacker_pos][f"seg{seg}"] += dmg
            op.execute_post_attack_buffs_for_attacker(mgr, attacker_pos, op_result)

        seg_damage = {f"seg{i}": 0 for i in range(6)}
        enemy_totals = {0: 0, 1: 0, 2: 0}
        total_damage = 0
        yang_damage = 0
        yin_damage = 0
        for d in all_details:
            dmg = int(d.get("damage_int", 0) or 0)
            seg = int(d.get("seg", -1))
            enemy_pos = int(d.get("enemy_pos", -1))
            if 0 <= seg < 6:
                seg_damage[f"seg{seg}"] += dmg
            if enemy_pos in enemy_totals:
                enemy_totals[enemy_pos] += dmg
            total_damage += dmg
            yy = d.get("yinyang")
            if yy == 1:
                yang_damage += dmg
            elif yy == 0:
                yin_damage += dmg

        yy_total = yang_damage + yin_damage
        if yy_total <= 0:
            yang_ratio = 0.0
            yin_ratio = 0.0
            yinyang_role = "阴阳均衡"
        else:
            yang_ratio = yang_damage / yy_total
            yin_ratio = yin_damage / yy_total
            if yang_ratio >= 0.7:
                yinyang_role = "主阳"
            elif yin_ratio >= 0.7:
                yinyang_role = "主阴"
            else:
                yinyang_role = "阴阳均衡"

        return {
            "seg_damage": seg_damage,
            "enemy_totals": enemy_totals,
            "ally_totals": ally_totals,
            "ally_seg_damage": ally_seg_damage,
            "total_damage": total_damage,
            "yang_damage_total": yang_damage,
            "yin_damage_total": yin_damage,
            "yang_ratio": yang_ratio,
            "yin_ratio": yin_ratio,
            "yinyang_role": yinyang_role,
            "details": all_details,
            "op_result": op_result,
            "attack_order": list(order),
            "attack_order_snapshot": snapshot,
            "skill_order": skill_order,
            "field_buffs": asdict(process_config.field_buffs),
            "selected_equipment": selected_equipment,
        }

    def evaluate_equipment_damage(self, char_id: int, attack_type: str, equipment_id: int, *, mode: str) -> float:
        enemy_slots = {
            0: EnemySlotConfig(
                enabled=True,
                character_id=1001,
                hp=99_999_999,
                yang_def=1000,
                yin_def=1000,
                barrier_count=9,
                quality=[1] * 9,
                tribe_text="",
                is_break_all=False,
                buffs=[],
            ),
            1: EnemySlotConfig(enabled=False),
            2: EnemySlotConfig(enabled=False),
        }
        equipment_ids = {key: 0 for key in EQUIPMENT_SLOT_KEYS}
        slot_key = self.get_equipment_slot_key(attack_type)
        if slot_key:
            equipment_ids[slot_key] = int(equipment_id or 0)
        spirit_level = 3 if str(attack_type) in {"1c", "2c", "5"} else 0
        ally_slots = {
            0: AllySlotConfig(
                enabled=True,
                character_id=int(char_id),
                initial_spirit=3.0,
                barrier_count=5,
                skill_order_text="",
                shield_open_count=0,
                attack_type=str(attack_type),
                spirit_level=spirit_level,
                target_enemy_pos=0,
                buffs=[],
                equipment_ids=equipment_ids,
            ),
            1: AllySlotConfig(enabled=False),
            2: AllySlotConfig(enabled=False),
        }
        result = self.run_single(enemy_slots, ally_slots, process_config=ProcessConfig(), equipment_mode=mode)
        return float(result.get("total_damage", 0) or 0.0)

    def _calc_enemy_seg_flag_counts(self, result: Dict[str, Any]) -> Dict[str, Any]:
        details = result.get("details", []) or []
        valid_segs = set(range(6))
        weak_segs_by_enemy = {0: set(), 1: set(), 2: set()}
        killer_segs_by_enemy = {0: set(), 1: set(), 2: set()}

        for d in details:
            try:
                seg = int(d.get("seg", -1))
                enemy_pos = int(d.get("enemy_pos", -1))
            except Exception:
                continue
            if seg not in valid_segs or enemy_pos not in (0, 1, 2):
                continue
            ei = d.get("element_info", {}) or {}
            hit = d.get("hit", {}) or {}
            if (ei.get("quality_value") == 0) or (ei.get("element_mode") == "advantage"):
                weak_segs_by_enemy[enemy_pos].add(seg)
            if str(hit.get("crit_mode", "")).lower() == "killer":
                killer_segs_by_enemy[enemy_pos].add(seg)

        return {
            "enemy_pos_0_weak_seg_count": len(weak_segs_by_enemy[0]),
            "enemy_pos_0_killer_seg_count": len(killer_segs_by_enemy[0]),
            "enemy_pos_1_weak_seg_count": len(weak_segs_by_enemy[1]),
            "enemy_pos_1_killer_seg_count": len(killer_segs_by_enemy[1]),
            "enemy_pos_2_weak_seg_count": len(weak_segs_by_enemy[2]),
            "enemy_pos_2_killer_seg_count": len(killer_segs_by_enemy[2]),
        }

    def run_batch_single_template(
        self,
        enemy_slots: Dict[int, EnemySlotConfig],
        ally_template: AllySlotConfig,
        process_config: Optional[ProcessConfig] = None,
    ) -> List[Dict[str, Any]]:
        process_config = process_config or ProcessConfig()
        rows: List[Dict[str, Any]] = []
        for cid in self.discover_all_ids():
            cfg0 = AllySlotConfig(**asdict(ally_template))
            cfg0.character_id = cid
            cfg0.equipment_ids = self.fill_missing_recommended_equipment_ids(cid, cfg0.equipment_ids)
            ally_slots = {0: cfg0, 1: AllySlotConfig(enabled=False), 2: AllySlotConfig(enabled=False)}
            try:
                meta = self.load_character_meta(cid)
                selected_equipment = self.get_equipment_for_attack_type(cfg0.equipment_ids, cfg0.attack_type) or {}
                result = self.run_single(enemy_slots, ally_slots, process_config=process_config)
                seg_flag_stats = self._calc_enemy_seg_flag_counts(result)
                row = {
                    "character_id": cid,
                    "character_name": meta.get("name", ""),
                    "name_cn": self.translate_character_name(meta.get("name", ""), cid),
                    "world_group": meta.get("world_group", ""),
                    "character_type": meta.get("type_label", ""),
                    "attack_type": cfg0.attack_type,
                    "equipment_id": int(selected_equipment.get("equipment_id", 0) or 0),
                    "equipment_name": str(selected_equipment.get("name", "") or ""),
                    "skill_order": cfg0.skill_order_text,
                    "spirit_level": cfg0.spirit_level,
                    "target_enemy_pos": cfg0.target_enemy_pos,
                    "seg0": result["seg_damage"].get("seg0", 0),
                    "seg1": result["seg_damage"].get("seg1", 0),
                    "seg2": result["seg_damage"].get("seg2", 0),
                    "seg3": result["seg_damage"].get("seg3", 0),
                    "seg4": result["seg_damage"].get("seg4", 0),
                    "seg5": result["seg_damage"].get("seg5", 0),
                    "enemy_pos_0_total": result["enemy_totals"].get(0, 0),
                    "enemy_pos_1_total": result["enemy_totals"].get(1, 0),
                    "enemy_pos_2_total": result["enemy_totals"].get(2, 0),
                    "overall_total": result.get("total_damage", 0),
                    "yang_damage_total": result.get("yang_damage_total", 0),
                    "yin_damage_total": result.get("yin_damage_total", 0),
                    "yang_ratio": result.get("yang_ratio", 0.0),
                    "yin_ratio": result.get("yin_ratio", 0.0),
                    "yinyang_role": result.get("yinyang_role", "阴阳均衡"),
                    "enemy_pos_0_weak_seg_count": seg_flag_stats["enemy_pos_0_weak_seg_count"],
                    "enemy_pos_0_killer_seg_count": seg_flag_stats["enemy_pos_0_killer_seg_count"],
                    "enemy_pos_1_weak_seg_count": seg_flag_stats["enemy_pos_1_weak_seg_count"],
                    "enemy_pos_1_killer_seg_count": seg_flag_stats["enemy_pos_1_killer_seg_count"],
                    "enemy_pos_2_weak_seg_count": seg_flag_stats["enemy_pos_2_weak_seg_count"],
                    "enemy_pos_2_killer_seg_count": seg_flag_stats["enemy_pos_2_killer_seg_count"],
                    "status": "ok",
                    "reason": "",
                }
            except Exception as e:
                row = {
                    "character_id": cid,
                    "character_name": "",
                    "name_cn": "",
                    "world_group": "",
                    "character_type": "",
                    "attack_type": ally_template.attack_type,
                    "equipment_id": 0,
                    "equipment_name": "",
                    "skill_order": ally_template.skill_order_text,
                    "spirit_level": ally_template.spirit_level,
                    "target_enemy_pos": ally_template.target_enemy_pos,
                    "seg0": 0,
                    "seg1": 0,
                    "seg2": 0,
                    "seg3": 0,
                    "seg4": 0,
                    "seg5": 0,
                    "enemy_pos_0_total": 0,
                    "enemy_pos_1_total": 0,
                    "enemy_pos_2_total": 0,
                    "overall_total": 0,
                    "yang_damage_total": 0,
                    "yin_damage_total": 0,
                    "yang_ratio": 0.0,
                    "yin_ratio": 0.0,
                    "yinyang_role": "阴阳均衡",
                    "enemy_pos_0_weak_seg_count": 0,
                    "enemy_pos_0_killer_seg_count": 0,
                    "enemy_pos_1_weak_seg_count": 0,
                    "enemy_pos_1_killer_seg_count": 0,
                    "enemy_pos_2_weak_seg_count": 0,
                    "enemy_pos_2_killer_seg_count": 0,
                    "status": "error",
                    "reason": repr(e),
                }
            rows.append(row)
        return rows
